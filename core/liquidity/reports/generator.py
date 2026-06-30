"""
Модуль отчётов — генератор.

Читает результаты расчётов из DWH/mart и формирует отчёты в трёх форматах:
  • Excel (.xlsx) — многолистовая книга с форматированием
  • PDF  (.pdf)   — табличный отчёт через reportlab
  • CSV  (.csv)   — плоский файл для выгрузки

Источники данных:
  • mart.liquidity_gap_view          — ГЭП-анализ (денормализованная витрина)
  • dwh.concentrationresult          — концентрация по контрагентам
  • dwh.gapcalculation               — мета-информация о расчёте

Отчёт регистрируется/обновляется в dwh.reporttask.

CLI:
    uv run report --date 2025-12-01
    uv run report --date 2025-12-01 --format excel
    uv run report --date 2025-12-01 --format pdf   --type gap
    uv run report --date 2025-12-01 --format csv   --type concentration
"""
from __future__ import annotations

import csv
import sys
import traceback
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Literal

import click
import pandas as pd
from sqlalchemy import text

from liquidity.config import app_settings
from liquidity.db import dwh_session
from liquidity.logger import get_logger, setup_logging

log = get_logger(__name__)

ReportFormat = Literal["excel", "pdf", "csv"]
ReportType   = Literal["gap", "concentration", "full"]

# ------------------------------------------------------------------
# Вспомогательные функции загрузки данных
# ------------------------------------------------------------------

def _load_gap_data(report_date: date, calculation_id: int | None = None) -> pd.DataFrame:
    """Загружает ГЭП-данные из mart.liquidity_gap_view."""
    with dwh_session() as session:
        params: dict = {"rd": report_date}
        extra_filter = ""
        if calculation_id is not None:
            extra_filter = "AND calculation_id = :cid"
            params["cid"] = calculation_id
        rows = session.execute(text(f"""
            SELECT
                bucket_code, bucket_name, sort_order,
                total_assets_rub, total_liabilities_rub,
                gap_rub, cumulative_gap_rub, gap_ratio_pct,
                calculation_id
            FROM mart.liquidity_gap_view
            WHERE report_date = :rd
              {extra_filter}
            ORDER BY sort_order
        """), params).fetchall()

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows, columns=[
        "bucket_code", "bucket_name", "sort_order",
        "total_assets_rub", "total_liabilities_rub",
        "gap_rub", "cumulative_gap_rub", "gap_ratio_pct",
        "calculation_id",
    ])
    # Decimal из PostgreSQL → float, чтобы pandas мог делать арифметику
    for col in ("total_assets_rub", "total_liabilities_rub",
                "gap_rub", "cumulative_gap_rub", "gap_ratio_pct"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    # Если несколько расчётов — берём последний
    if df["calculation_id"].nunique() > 1:
        latest_cid = df["calculation_id"].max()
        df = df[df["calculation_id"] == latest_cid].copy()
    return df


def _load_concentration_data(
    report_date: date, calculation_id: int | None = None
) -> pd.DataFrame:
    """Загружает данные концентрации из dwh.concentrationresult."""
    with dwh_session() as session:
        params: dict = {"rd": report_date}
        extra_filter = ""
        if calculation_id is not None:
            extra_filter = "AND cr.calculation_id = :cid"
            params["cid"] = calculation_id
        rows = session.execute(text(f"""
            SELECT
                cr.calculation_id,
                cr.category,
                cp.code           AS counterparty_code,
                cp.short_name     AS counterparty_name,
                cp.counterparty_type,
                tb.code           AS bucket_code,
                tb.name           AS bucket_name,
                cr.amount_rub,
                cr.share_pct
            FROM dwh.concentrationresult cr
            JOIN dwh.counterpartyref cp ON cp.id = cr.counterparty_id
            JOIN dwh.timebucket tb      ON tb.id = cr.timebucket_id
            WHERE cr.report_date = :rd
              {extra_filter}
            ORDER BY cr.category, cr.share_pct DESC
        """), params).fetchall()

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows, columns=[
        "calculation_id", "category",
        "counterparty_code", "counterparty_name", "counterparty_type",
        "bucket_code", "bucket_name",
        "amount_rub", "share_pct",
    ])
    # Decimal из PostgreSQL → float
    df["amount_rub"] = pd.to_numeric(df["amount_rub"], errors="coerce")
    df["share_pct"]  = pd.to_numeric(df["share_pct"],  errors="coerce")
    if df["calculation_id"].nunique() > 1:
        latest_cid = df["calculation_id"].max()
        df = df[df["calculation_id"] == latest_cid].copy()
    return df


def _get_latest_calculation(report_date: date) -> int | None:
    """Возвращает id последнего успешного расчёта для даты."""
    with dwh_session() as session:
        row = session.execute(text("""
            SELECT id FROM dwh.gapcalculation
            WHERE report_date = :rd AND status = 'success'
            ORDER BY finished_at DESC NULLS LAST
            LIMIT 1
        """), {"rd": report_date}).fetchone()
    return row.id if row else None


# ------------------------------------------------------------------
# Excel-генератор
# ------------------------------------------------------------------

def _generate_excel(
    report_date: date,
    gap_df: pd.DataFrame,
    conc_df: pd.DataFrame,
    report_type: ReportType,
) -> bytes:
    """
    Формирует многолистовую Excel-книгу.
    Лист 1 — ГЭП-анализ
    Лист 2 — Концентрация активов
    Лист 3 — Концентрация обязательств
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # удаляем дефолтный пустой лист

    # --- Стили ---
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
    SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
    NEG_FILL      = PatternFill("solid", fgColor="FCE4D6")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
    SUBHDR_FONT   = Font(bold=True, color="FFFFFF", size=9)
    NORMAL_FONT   = Font(size=9)
    BOLD_FONT     = Font(bold=True, size=9)
    TITLE_FONT    = Font(bold=True, size=12, color="1F4E79")

    thin = Side(style="thin", color="BFBFBF")
    med  = Side(style="medium", color="1F4E79")
    BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
    BORDER_MED  = Border(left=med, right=med, top=med, bottom=med)

    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT  = Alignment(horizontal="right",  vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    RUB_FMT  = '#,##0.00 ₽'
    PCT_FMT  = '0.00%'
    PCT2_FMT = '0.000%'

    def _set_header_row(ws, row_idx: int, headers: list[tuple[str, int]]) -> None:
        """Записывает строку заголовка с шириной колонок."""
        for col_idx, (title, width) in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=title)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.border = BORDER_THIN
            cell.alignment = CENTER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _apply_data_row(ws, row_idx: int, values: list, fills: list | None = None) -> None:
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = NORMAL_FONT
            cell.border = BORDER_THIN
            cell.alignment = RIGHT if isinstance(val, (int, float)) else LEFT
            if fills and fills[col_idx - 1]:
                cell.fill = fills[col_idx - 1]

    # ================================================================
    # Лист 1: ГЭП-анализ
    # ================================================================
    if report_type in ("gap", "full") and not gap_df.empty:
        ws = wb.create_sheet("ГЭП-анализ")
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A4"

        # Заголовок
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value     = f"Анализ разрывов ликвидности (ГЭП-анализ) | Дата отчёта: {report_date}"
        title_cell.font      = TITLE_FONT
        title_cell.alignment = CENTER

        # Подзаголовок
        ws.merge_cells("A2:I2")
        sub = ws["A2"]
        sub.value     = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Единица измерения: руб."
        sub.font      = Font(italic=True, size=9, color="595959")
        sub.alignment = CENTER

        headers = [
            ("Корзина (код)",       14),
            ("Корзина (название)",  22),
            ("Активы, руб.",        18),
            ("Обязательства, руб.", 20),
            ("ГЭП, руб.",           18),
            ("Накопл. ГЭП, руб.",   20),
            ("ГЭП/Обяз., %",        14),
            ("Статус",               12),
        ]
        _set_header_row(ws, 3, headers)

        total_assets = 0.0
        total_liabs  = 0.0

        for i, r in enumerate(gap_df.itertuples(), start=4):
            gap_val  = float(r.gap_rub) if r.gap_rub is not None else 0.0
            cum_gap  = float(r.cumulative_gap_rub) if r.cumulative_gap_rub is not None else 0.0
            ratio    = float(r.gap_ratio_pct) / 100.0 if r.gap_ratio_pct is not None else None
            assets   = float(r.total_assets_rub) if r.total_assets_rub is not None else 0.0
            liabs    = float(r.total_liabilities_rub) if r.total_liabilities_rub is not None else 0.0

            status = "✓ Профицит" if gap_val >= 0 else "✗ Дефицит"
            row_fill = ALT_FILL if i % 2 == 0 else None
            neg_fill = NEG_FILL if gap_val < 0 else row_fill

            values = [r.bucket_code, r.bucket_name, assets, liabs, gap_val, cum_gap, ratio, status]
            fills  = [row_fill] * 4 + [neg_fill, neg_fill, row_fill, neg_fill if gap_val < 0 else row_fill]
            _apply_data_row(ws, i, values, fills)

            # Форматирование числовых ячеек
            ws.cell(i, 3).number_format = RUB_FMT
            ws.cell(i, 4).number_format = RUB_FMT
            ws.cell(i, 5).number_format = RUB_FMT
            ws.cell(i, 6).number_format = RUB_FMT
            ws.cell(i, 7).number_format = PCT_FMT

            total_assets += assets
            total_liabs  += liabs

        # Итого
        total_row = len(gap_df) + 4
        ws.merge_cells(f"A{total_row}:B{total_row}")
        ws.cell(total_row, 1).value = "ИТОГО"
        ws.cell(total_row, 1).font  = BOLD_FONT
        ws.cell(total_row, 1).fill  = SUBHDR_FILL
        ws.cell(total_row, 1).font  = SUBHDR_FONT
        ws.cell(total_row, 3).value = total_assets
        ws.cell(total_row, 3).number_format = RUB_FMT
        ws.cell(total_row, 3).font = BOLD_FONT
        ws.cell(total_row, 4).value = total_liabs
        ws.cell(total_row, 4).number_format = RUB_FMT
        ws.cell(total_row, 4).font = BOLD_FONT
        ws.cell(total_row, 5).value = total_assets - total_liabs
        ws.cell(total_row, 5).number_format = RUB_FMT
        ws.cell(total_row, 5).font = BOLD_FONT
        for col in range(1, 9):
            ws.cell(total_row, col).border = BORDER_THIN
            if ws.cell(total_row, col).fill == PatternFill():
                ws.cell(total_row, col).fill = SUBHDR_FILL

    # ================================================================
    # Лист 2/3: Концентрация (активы и обязательства)
    # ================================================================
    if report_type in ("concentration", "full") and not conc_df.empty:
        for category, sheet_title in [("asset", "Концентрация — Активы"),
                                       ("liability", "Концентрация — Обязательства")]:
            cat_df = conc_df[conc_df["category"] == category].copy()
            if cat_df.empty:
                continue

            ws = wb.create_sheet(sheet_title)
            ws.freeze_panes = "A4"

            ws.merge_cells("A1:H1")
            ws["A1"].value     = f"{sheet_title} | Дата отчёта: {report_date}"
            ws["A1"].font      = TITLE_FONT
            ws["A1"].alignment = CENTER

            ws.merge_cells("A2:H2")
            ws["A2"].value     = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')} | Топ контрагентов по доле в портфеле"
            ws["A2"].font      = Font(italic=True, size=9, color="595959")
            ws["A2"].alignment = CENTER

            headers = [
                ("Код контрагента", 16),
                ("Наименование",    28),
                ("Тип",             14),
                ("Корзина",         20),
                ("Сумма, руб.",     18),
                ("Доля, %",         12),
                ("Доля (кумул.)",   14),
                ("Топ-3?",           8),
            ]
            _set_header_row(ws, 3, headers)

            # Группируем по контрагенту для совокупной доли
            cp_totals = (
                cat_df.groupby("counterparty_code")["amount_rub"].sum()
                      .sort_values(ascending=False)
            )
            total_portfolio = cp_totals.sum()
            cp_shares = (cp_totals / total_portfolio * 100).round(3) if total_portfolio > 0 else cp_totals * 0

            cumulative = 0.0
            for i, r in enumerate(cat_df.itertuples(), start=4):
                cp_share = float(cp_shares.get(r.counterparty_code, 0))
                cumulative += float(r.share_pct)
                top3 = "★" if cp_share >= cp_shares.iloc[2] and len(cp_shares) >= 3 else ""

                row_fill = ALT_FILL if i % 2 == 0 else None
                values = [
                    r.counterparty_code,
                    r.counterparty_name,
                    r.counterparty_type,
                    r.bucket_name,
                    float(r.amount_rub),
                    float(r.share_pct) / 100.0,
                    min(cumulative, 100.0) / 100.0,
                    top3,
                ]
                fills = [row_fill] * 8
                _apply_data_row(ws, i, values, fills)

                ws.cell(i, 5).number_format = RUB_FMT
                ws.cell(i, 6).number_format = PCT2_FMT
                ws.cell(i, 7).number_format = PCT2_FMT
                ws.cell(i, 8).alignment = CENTER

    # Сохраняем в байты
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------
# PDF-генератор — поддержка кириллицы
# ------------------------------------------------------------------

def _register_cyrillic_font() -> tuple[str, str]:
    """
    Регистрирует TTF-шрифт с поддержкой кириллицы в ReportLab.
    Возвращает (font_regular, font_bold) — имена для использования в стилях.

    Порядок поиска:
      macOS → Arial / DejaVuSans
      Linux → DejaVuSans
      Windows → Arial
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        # macOS
        ("/Library/Fonts/Arial.ttf",                              "/Library/Fonts/Arial Bold.ttf"),
        ("/System/Library/Fonts/Supplemental/Arial.ttf",          "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        # macOS — DejaVu (устанавливается вместе со многими пакетами)
        ("/opt/homebrew/share/fonts/dejavu/DejaVuSans.ttf",       "/opt/homebrew/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
        # Linux
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",       "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/dejavu/DejaVuSans.ttf",                "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
        # Windows
        ("C:/Windows/Fonts/arial.ttf",                            "C:/Windows/Fonts/arialbd.ttf"),
    ]

    for reg_path, bold_path in candidates:
        if Path(reg_path).exists():
            try:
                pdfmetrics.registerFont(TTFont("CyrFont",     reg_path))
                if Path(bold_path).exists():
                    pdfmetrics.registerFont(TTFont("CyrFontBd", bold_path))
                else:
                    pdfmetrics.registerFont(TTFont("CyrFontBd", reg_path))
                log.info("pdf.font_registered", path=reg_path)
                return "CyrFont", "CyrFontBd"
            except Exception as e:
                log.warning("pdf.font_register_failed", path=reg_path, error=str(e))
                continue

    # Запасной вариант: Helvetica без кириллицы (лучше, чем падение)
    log.warning("pdf.no_cyrillic_font_found")
    return "Helvetica", "Helvetica-Bold"


def _generate_pdf(
    report_date: date,
    gap_df: pd.DataFrame,
    conc_df: pd.DataFrame,
    report_type: ReportType,
) -> bytes:
    """Формирует PDF-отчёт через reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    font_reg, font_bold = _register_cyrillic_font()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title",
        parent=styles["Heading1"],
        fontName=font_bold,
        fontSize=14,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "subtitle",
        parent=styles["Normal"],
        fontName=font_reg,
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "section",
        parent=styles["Heading2"],
        fontName=font_bold,
        fontSize=11,
        textColor=colors.HexColor("#2E75B6"),
        spaceBefore=12,
        spaceAfter=6,
    )

    DARK_BLUE  = colors.HexColor("#1F4E79")
    MED_BLUE   = colors.HexColor("#2E75B6")
    LIGHT_BLUE = colors.HexColor("#EBF3FB")
    ORANGE     = colors.HexColor("#FCE4D6")

    story = []

    story.append(Paragraph(
        f"Отчёт о структурной ликвидности банка | Дата: {report_date}",
        title_style,
    ))
    story.append(Paragraph(
        f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')} | "
        "Система автоматизированного формирования отчётности по структурной ликвидности",
        subtitle_style,
    ))

    def _fmt_rub(val) -> str:
        if val is None:
            return "—"
        v = float(val)
        import math
        if math.isnan(v) or math.isinf(v):
            return "—"
        return f"{v:,.2f} руб."

    def _fmt_pct(val) -> str:
        if val is None:
            return "—"
        import math
        v = float(val)
        if math.isnan(v) or math.isinf(v):
            return "—"
        return f"{v:.2f}%"

    # ================================================
    # ГЭП-анализ
    # ================================================
    if report_type in ("gap", "full") and not gap_df.empty:
        story.append(Paragraph("1. Анализ разрывов ликвидности (ГЭП-анализ)", section_style))

        table_data = [[
            "Корзина", "Название", "Активы, руб.", "Обязательства, руб.",
            "ГЭП, руб.", "Накопл. ГЭП", "ГЭП/%",
        ]]
        for r in gap_df.itertuples():
            gap_val = float(r.gap_rub) if r.gap_rub is not None else 0.0
            table_data.append([
                r.bucket_code,
                r.bucket_name,
                _fmt_rub(r.total_assets_rub),
                _fmt_rub(r.total_liabilities_rub),
                _fmt_rub(gap_val),
                _fmt_rub(r.cumulative_gap_rub),
                _fmt_pct(r.gap_ratio_pct),
            ])

        # Итого
        total_a = gap_df["total_assets_rub"].sum()
        total_l = gap_df["total_liabilities_rub"].sum()
        table_data.append([
            "ИТОГО", "",
            _fmt_rub(total_a), _fmt_rub(total_l),
            _fmt_rub(total_a - total_l), "", "",
        ])

        col_widths = [3*cm, 5.5*cm, 4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm, 3*cm]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        ts = TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  DARK_BLUE),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",     (0, 0), (-1, 0),  font_bold),
            ("FONTSIZE",     (0, 0), (-1, 0),  8),
            ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, LIGHT_BLUE]),
            ("BACKGROUND",   (0, -1), (-1, -1), MED_BLUE),
            ("TEXTCOLOR",    (0, -1), (-1, -1), colors.white),
            ("FONTNAME",     (0, -1), (-1, -1), font_bold),
            ("FONTNAME",     (0, 1),  (-1, -2), font_reg),
            ("FONTSIZE",     (0, 1), (-1, -1),  8),
            ("ALIGN",        (2, 1), (-1, -1),  "RIGHT"),
            ("GRID",         (0, 0), (-1, -1),  0.5, colors.HexColor("#BFBFBF")),
            ("TOPPADDING",   (0, 0), (-1, -1),  3),
            ("BOTTOMPADDING",(0, 0), (-1, -1),  3),
        ])

        # Красим дефицитные строки
        for row_idx, r in enumerate(gap_df.itertuples(), start=1):
            gap_val = float(r.gap_rub) if r.gap_rub is not None else 0.0
            if gap_val < 0:
                ts.add("BACKGROUND", (4, row_idx), (5, row_idx), ORANGE)
                ts.add("TEXTCOLOR",  (4, row_idx), (5, row_idx), colors.HexColor("#C00000"))

        t.setStyle(ts)
        story.append(t)

    # ================================================
    # Концентрация
    # ================================================
    if report_type in ("concentration", "full") and not conc_df.empty:
        for category, label in [("asset", "активов"), ("liability", "обязательств")]:
            cat_df = conc_df[conc_df["category"] == category].copy()
            if cat_df.empty:
                continue

            section_num = 2 if report_type == "full" and category == "asset" else \
                          3 if report_type == "full" else 1
            story.append(Paragraph(
                f"{section_num}. Концентрация {label} по контрагентам",
                section_style,
            ))

            # Топ-15 по доле
            top_df = (
                cat_df.groupby(["counterparty_code", "counterparty_name", "counterparty_type"])
                ["amount_rub"].sum()
                .sort_values(ascending=False)
                .head(15)
                .reset_index()
            )
            total = float(top_df["amount_rub"].sum())

            table_data = [["#", "Код", "Наименование", "Тип", "Сумма, руб.", "Доля, %"]]
            cumulative = 0.0
            for rank, row in enumerate(top_df.itertuples(), start=1):
                share = float(row.amount_rub) / total * 100 if total > 0 else 0
                cumulative += share
                table_data.append([
                    str(rank),
                    row.counterparty_code,
                    row.counterparty_name,
                    row.counterparty_type,
                    _fmt_rub(row.amount_rub),
                    f"{share:.3f}%",
                ])

            col_widths = [1*cm, 3*cm, 7*cm, 3.5*cm, 5*cm, 3*cm]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, 0),  DARK_BLUE),
                ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
                ("FONTNAME",     (0, 0), (-1, 0),  font_bold),
                ("FONTSIZE",     (0, 0), (-1, 0),  8),
                ("ALIGN",        (0, 0), (-1, 0),  "CENTER"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BLUE]),
                ("FONTNAME",     (0, 1), (-1, -1),  font_reg),
                ("FONTSIZE",     (0, 1), (-1, -1),  8),
                ("ALIGN",        (0, 1), (1, -1),  "CENTER"),
                ("ALIGN",        (4, 1), (-1, -1),  "RIGHT"),
                ("GRID",         (0, 0), (-1, -1),  0.5, colors.HexColor("#BFBFBF")),
                ("TOPPADDING",   (0, 0), (-1, -1),  3),
                ("BOTTOMPADDING",(0, 0), (-1, -1),  3),
            ]))
            story.append(t)
            story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    return buf.getvalue()


# ------------------------------------------------------------------
# CSV-генератор
# ------------------------------------------------------------------

def _generate_csv(
    report_date: date,
    gap_df: pd.DataFrame,
    conc_df: pd.DataFrame,
    report_type: ReportType,
    out_dir: Path,
    base_name: str,
) -> list[Path]:
    """Сохраняет CSV-файлы (один или несколько). Возвращает список путей."""
    paths: list[Path] = []

    if report_type in ("gap", "full") and not gap_df.empty:
        p = out_dir / f"{base_name}_gap.csv"
        gap_df.to_csv(p, index=False, encoding="utf-8-sig",
                      quoting=csv.QUOTE_NONNUMERIC)
        paths.append(p)

    if report_type in ("concentration", "full") and not conc_df.empty:
        p = out_dir / f"{base_name}_concentration.csv"
        conc_df.to_csv(p, index=False, encoding="utf-8-sig",
                       quoting=csv.QUOTE_NONNUMERIC)
        paths.append(p)

    return paths


# ------------------------------------------------------------------
# ReportGenerator — публичный класс
# ------------------------------------------------------------------

class ReportGenerator:
    """
    Генерирует отчёт по ликвидности и регистрирует задачу в dwh.reporttask.
    """

    def __init__(
        self,
        report_date: date,
        report_type:   ReportType   = "full",
        report_format: ReportFormat = "excel",
        output_dir: Path | None = None,
        initiated_by: int | None = None,
    ) -> None:
        self.report_date   = report_date
        self.report_type   = report_type
        self.report_format = report_format
        self.output_dir    = output_dir or app_settings.reports_output_dir
        self.initiated_by  = initiated_by
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------

    def run(self) -> list[Path]:
        """Полный цикл: загрузка данных → генерация → сохранение → журнал."""
        task_id = self._create_task_record()
        log.info("report_generator.start",
                 report_date=str(self.report_date),
                 report_type=self.report_type,
                 report_format=self.report_format)
        try:
            paths = self._generate()
            self._update_task_record(task_id, "success",
                                     file_path="; ".join(str(p) for p in paths))
            log.info("report_generator.done",
                     files=[str(p) for p in paths])
            return paths
        except Exception as exc:
            self._update_task_record(task_id, "failed", error=str(exc))
            log.error("report_generator.failed", exc_info=exc)
            raise

    def _generate(self) -> list[Path]:
        calc_id = _get_latest_calculation(self.report_date)
        log.info("report_generator.loading_data",
                 calculation_id=calc_id,
                 report_date=str(self.report_date))

        gap_df  = _load_gap_data(self.report_date, calc_id) \
                  if self.report_type in ("gap", "full") else pd.DataFrame()
        conc_df = _load_concentration_data(self.report_date, calc_id) \
                  if self.report_type in ("concentration", "full") else pd.DataFrame()

        if gap_df.empty and conc_df.empty:
            raise RuntimeError(
                f"Нет данных за {self.report_date}. "
            )

        base_name = (
            f"liquidity_report_{self.report_date}_{self.report_type}"
        )

        if self.report_format == "excel":
            data = _generate_excel(self.report_date, gap_df, conc_df, self.report_type)
            out  = self.output_dir / f"{base_name}.xlsx"
            out.write_bytes(data)
            return [out]

        elif self.report_format == "pdf":
            data = _generate_pdf(self.report_date, gap_df, conc_df, self.report_type)
            out  = self.output_dir / f"{base_name}.pdf"
            out.write_bytes(data)
            return [out]

        elif self.report_format == "csv":
            return _generate_csv(
                self.report_date, gap_df, conc_df,
                self.report_type, self.output_dir, base_name,
            )

        raise ValueError(f"Неизвестный формат: {self.report_format}")

    # ------------------------------------------------------------------
    # Журнал в dwh.reporttask
    # ------------------------------------------------------------------

    # Маппинг CLI-формата → код в БД (CHECK constraint: 'xlsx', 'pdf', 'csv')
    _DB_FORMAT = {"excel": "xlsx", "pdf": "pdf", "csv": "csv"}

    def _create_task_record(self) -> int:
        db_format = self._DB_FORMAT.get(self.report_format, self.report_format)
        ext = "xlsx" if self.report_format == "excel" else self.report_format
        with dwh_session() as session:
            row = session.execute(text("""
                INSERT INTO dwh.reporttask
                    (report_date, report_type, report_format, status,
                     report_name, initiated_by)
                VALUES (:rd, :rt, :rf, 'running', :rn, :uid)
                RETURNING id
            """), {
                "rd":  self.report_date,
                "rt":  self.report_type,
                "rf":  db_format,
                "rn":  f"liquidity_report_{self.report_date}_{self.report_type}.{ext}",
                "uid": self.initiated_by,
            }).fetchone()
        return row.id

    def _update_task_record(
        self, task_id: int, status: str,
        file_path: str | None = None,
        error: str | None = None,
    ) -> None:
        with dwh_session() as session:
            session.execute(text("""
                UPDATE dwh.reporttask
                SET status        = :status,
                    finished_at   = NOW(),
                    file_path     = :fp,
                    error_message = :err
                WHERE id = :tid
            """), {"status": status, "fp": file_path, "err": error, "tid": task_id})


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

@click.command(name="report")
@click.option(
    "--date", "report_date_str",
    required=True, metavar="YYYY-MM-DD",
    help="Дата отчёта.",
)
@click.option(
    "--format", "report_format",
    type=click.Choice(["excel", "pdf", "csv"]),
    default="excel", show_default=True,
    help="Формат выходного файла.",
)
@click.option(
    "--type", "report_type",
    type=click.Choice(["gap", "concentration", "full"]),
    default="full", show_default=True,
    help="Тип отчёта.",
)
@click.option(
    "--output-dir", "output_dir_str",
    default=None,
    help="Каталог для сохранения файла (по умолчанию из конфига).",
)
@click.option(
    "--user-id", "user_id",
    default=None, type=int,
    help="ID пользователя.",
)
def cli(
    report_date_str: str,
    report_format: str,
    report_type: str,
    output_dir_str: str | None,
    user_id: int | None,
) -> None:
    """Генерация отчёта о структурной ликвидности."""
    setup_logging()

    try:
        report_date = date.fromisoformat(report_date_str)
    except ValueError:
        click.echo(f"Ошибка: неверный формат даты '{report_date_str}'. Используй YYYY-MM-DD.")
        sys.exit(1)

    output_dir = Path(output_dir_str) if output_dir_str else None

    click.echo(f"Генерация отчёта: type={report_type}, format={report_format}, date={report_date}")
    generator = ReportGenerator(
        report_date=report_date,
        report_type=report_type,
        report_format=report_format,
        output_dir=output_dir,
        initiated_by=user_id,
    )

    try:
        paths = generator.run()
        for p in paths:
            click.echo(f"  Сохранено: {p}")
        sys.exit(0)
    except Exception:
        click.echo("\n--- TRACEBACK ---", err=True)
        click.echo(traceback.format_exc(), err=True)
        sys.exit(1)


if __name__ == "__main__":
    cli()
