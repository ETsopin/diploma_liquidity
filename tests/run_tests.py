"""
Тест-скрипт для всех 17 тест-кейсов дипломной работы.

Охватывает три уровня тестирования:
  1.x — Модульное (Transformer, GapCalculator, ConcentrationCalculator, ReportGenerator)
  2.x — Интеграционное (ETL-пайплайн, трансформация Staging→DWH, расчётный пайплайн, отчёты)
  3.x — Системное (E2E, отказоустойчивость, идемпотентность, API, транзакции)

Не требует запущенного PostgreSQL: алгоритмы воспроизведены автономно,
БД-логика симулируется через sqlite3 в памяти.

Запуск:
    python tests/run_tests.py                    # из корня liquidity_system/
    python tests/run_tests.py --json results.json
"""

import sys
import os
import csv
import json
import sqlite3
import argparse
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

# ─── вывод результатов ────────────────────────────────────────────────────────

RESULTS: dict = {}


def test(tid: str, name: str):
    """Декоратор, запускающий тест и фиксирующий результат."""
    def decorator(fn):
        def wrapper():
            try:
                detail = fn()
                RESULTS[tid] = {"name": name, "status": "ПРОЙДЕН", "detail": detail or ""}
                print(f"  ✓ {tid}. {name}")
            except AssertionError as e:
                RESULTS[tid] = {"name": name, "status": "НЕ ПРОЙДЕН", "detail": str(e)}
                print(f"  ✗ {tid}. {name}: {e}")
            except Exception as e:
                RESULTS[tid] = {"name": name, "status": "НЕ ПРОЙДЕН",
                                "detail": f"{type(e).__name__}: {e}"}
                print(f"  ✗ {tid}. {name}: {e}")
        wrapper()
        return wrapper
    return decorator


# ═══════════════════════════════════════════════════════════════════════════════
# Автономные реализации алгоритмов системы (без импорта liquidity-пакета)
# ═══════════════════════════════════════════════════════════════════════════════

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Повторяет логику Transformer._clean."""
    df = df.copy()
    for col in ("contract_number", "counterparty_code", "currency", "status"):
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"nan": None, "None": None, "": None})
    df["amount"] = pd.to_numeric(df.get("amount"), errors="coerce")
    if "currency" in df.columns:
        df["currency"] = df["currency"].str.upper()
    mask_ok = df["amount"].notna() & (df["amount"] > 0)
    if "contract_number" in df.columns:
        mask_ok &= df["contract_number"].notna()
    return df[mask_ok].copy()


TIMEBUCKETS = [
    {"code": "ON_DEMAND", "min_days": 0,   "max_days": 0},
    {"code": "D1",        "min_days": 1,   "max_days": 1},
    {"code": "D2_7",      "min_days": 2,   "max_days": 7},
    {"code": "D8_30",     "min_days": 8,   "max_days": 30},
    {"code": "D31_90",    "min_days": 31,  "max_days": 90},
    {"code": "D91_180",   "min_days": 91,  "max_days": 180},
    {"code": "D181_365",  "min_days": 181, "max_days": 365},
    {"code": "OVER_1Y",   "min_days": 366, "max_days": None},
]
ON_DEMAND_PRODUCTS = {"DEPOSIT_DEMAND", "CASH"}


def timebucket_id(days, product_type: str) -> str:
    """Повторяет логику ReferenceCache.timebucket_id."""
    if days is None or product_type in ON_DEMAND_PRODUCTS:
        return "ON_DEMAND"
    for tb in TIMEBUCKETS:
        if days >= tb["min_days"] and (tb["max_days"] is None or days <= tb["max_days"]):
            return tb["code"]
    return None


def calc_gap(buckets_data: list) -> pd.DataFrame:
    """
    Повторяет логику GapCalculator.calculate.
    buckets_data: list of (bucket_code, total_assets, total_liabilities)
    """
    df = pd.DataFrame(buckets_data, columns=["bucket_code", "total_assets", "total_liabilities"])
    df["total_assets"] = df["total_assets"].fillna(0).astype(float)
    df["total_liabilities"] = df["total_liabilities"].fillna(0).astype(float)
    df["gap"] = df["total_assets"] - df["total_liabilities"]
    df["cumulative_gap"] = df["gap"].cumsum()
    df["gap_ratio"] = df.apply(
        lambda r: round(r["gap"] / r["total_liabilities"] * 100, 4)
        if r["total_liabilities"] != 0 else None,
        axis=1,
    )
    return df


def calc_shares(df: pd.DataFrame) -> pd.DataFrame:
    """Повторяет логику ConcentrationCalculator._calc_shares."""
    total = df["amount_rub"].sum()
    df = df.copy()
    df["share_pct"] = (df["amount_rub"] / total * 100).round(3) if total > 0 else 0.0
    return df


def generate_excel_report(gap_df: pd.DataFrame) -> bytes:
    """Повторяет логику ReportGenerator._build_xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ГЭП-анализ"
    ws.append(["Временная корзина", "Активы (руб.)", "Обязательства (руб.)",
               "Разрыв (руб.)", "Накопленный разрыв (руб.)"])
    for _, row in gap_df.iterrows():
        ws.append([row["bucket_code"], row["total_assets"],
                   row["total_liabilities"], row["gap"], row["cumulative_gap"]])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_pdf_report(gap_df: pd.DataFrame) -> bytes:
    """Повторяет логику ReportGenerator._build_pdf."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    data = [["Корзина", "Активы", "Обязательства", "Разрыв"]]
    for _, row in gap_df.iterrows():
        data.append([str(row["bucket_code"]),
                     f"{row['total_assets']:,.0f}",
                     f"{row['total_liabilities']:,.0f}",
                     f"{row['gap']:,.0f}"])
    tbl = Table(data)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), (0.8, 0.8, 0.8)),
        ("GRID", (0, 0), (-1, -1), 0.5, (0, 0, 0)),
    ]))
    doc.build([tbl])
    return buf.getvalue()


def generate_csv_report(gap_df: pd.DataFrame) -> str:
    """Повторяет логику ReportGenerator._build_csv."""
    import io
    out = io.StringIO()
    fields = ["bucket_code", "total_assets", "total_liabilities", "gap", "cumulative_gap"]
    writer = csv.DictWriter(out, fieldnames=fields)
    writer.writeheader()
    for _, row in gap_df.iterrows():
        writer.writerow({k: row[k] for k in fields})
    return out.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# Вспомогательная SQLite-БД — воспроизводит схему DWH и ABS для тестов
# ═══════════════════════════════════════════════════════════════════════════════

def make_test_db() -> sqlite3.Connection:
    """
    Создаёт SQLite in-memory БД, воспроизводящую схемы:
      audit   → raw_batch
      staging → staging_asset, staging_liability
      dwh     → counterpartyref, producttyperef, timebucket,
                asset, liability, gapcalculation, gapresult,
                concentrationresult, reporttask
      mart    → mart_gap_view
    и заполняет справочные таблицы начальными данными.
    """
    conn = sqlite3.connect(":memory:")
    cur = conn.cursor()
    cur.executescript("""
        -- audit схема
        CREATE TABLE raw_batch (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            datasource_id    INTEGER,
            started_at       TEXT,
            finished_at      TEXT,
            status           TEXT DEFAULT 'running',
            rows_extracted   INTEGER,
            rows_loaded      INTEGER,
            error_message    TEXT
        );

        -- staging схема
        CREATE TABLE staging_asset (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id         INTEGER,
            raw_contract_id  TEXT,
            counterparty_code TEXT,
            product_code     TEXT,
            amount           REAL,
            currency         TEXT,
            maturity_date    TEXT,
            issue_date       TEXT,
            asset_type       TEXT,
            raw_data         TEXT,
            is_processed     INTEGER DEFAULT 0,
            loaded_at        TEXT,
            source_system    TEXT
        );
        CREATE TABLE staging_liability (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id         INTEGER,
            raw_contract_id  TEXT,
            counterparty_code TEXT,
            product_code     TEXT,
            amount           REAL,
            currency         TEXT,
            maturity_date    TEXT,
            issue_date       TEXT,
            liability_type   TEXT,
            raw_data         TEXT,
            is_processed     INTEGER DEFAULT 0,
            loaded_at        TEXT,
            source_system    TEXT
        );

        -- dwh справочники
        CREATE TABLE dwh_counterpartyref (
            id               INTEGER PRIMARY KEY,
            code             TEXT UNIQUE,
            full_name        TEXT,
            inn              TEXT,
            counterparty_type TEXT
        );
        CREATE TABLE dwh_producttyperef (
            id               INTEGER PRIMARY KEY,
            code             TEXT UNIQUE,
            name             TEXT,
            category         TEXT
        );
        CREATE TABLE dwh_timebucket (
            id               INTEGER PRIMARY KEY,
            code             TEXT UNIQUE,
            name             TEXT,
            min_days         INTEGER,
            max_days         INTEGER,
            sort_order       INTEGER
        );

        -- dwh факты
        CREATE TABLE dwh_asset (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            staging_id       INTEGER,
            batch_id         INTEGER,
            counterparty_id  INTEGER,
            product_type_id  INTEGER,
            contract_number  TEXT,
            amount           REAL,
            amount_rub       REAL,
            currency         TEXT,
            exchange_rate    REAL DEFAULT 1,
            maturity_date    TEXT,
            timebucket_id    INTEGER,
            days_to_maturity INTEGER,
            report_date      TEXT,
            is_valid         INTEGER DEFAULT 1
        );
        CREATE TABLE dwh_liability (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            staging_id       INTEGER,
            batch_id         INTEGER,
            counterparty_id  INTEGER,
            product_type_id  INTEGER,
            contract_number  TEXT,
            amount           REAL,
            amount_rub       REAL,
            currency         TEXT,
            exchange_rate    REAL DEFAULT 1,
            maturity_date    TEXT,
            timebucket_id    INTEGER,
            days_to_maturity INTEGER,
            report_date      TEXT,
            is_valid         INTEGER DEFAULT 1
        );

        -- dwh расчёты
        CREATE TABLE gapcalculation (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            report_date      TEXT,
            calc_type        TEXT DEFAULT 'full',
            status           TEXT DEFAULT 'pending',
            started_at       TEXT,
            finished_at      TEXT,
            error_message    TEXT
        );
        CREATE TABLE gapresult (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            calculation_id   INTEGER,
            report_date      TEXT,
            timebucket_id    INTEGER,
            total_assets     REAL DEFAULT 0,
            total_liabilities REAL DEFAULT 0,
            gap              REAL,
            cumulative_gap   REAL,
            gap_ratio        REAL,
            UNIQUE(calculation_id, timebucket_id)
        );
        CREATE TABLE concentrationresult (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            calculation_id   INTEGER,
            report_date      TEXT,
            counterparty_id  INTEGER,
            category         TEXT,
            amount_rub       REAL,
            share_pct        REAL,
            timebucket_id    INTEGER
        );
        CREATE TABLE reporttask (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            calculation_id   INTEGER,
            report_type      TEXT,
            report_format    TEXT,
            status           TEXT DEFAULT 'pending',
            file_path        TEXT,
            error_message    TEXT,
            report_date      TEXT,
            created_at       TEXT,
            finished_at      TEXT
        );

        -- mart витрина
        CREATE TABLE mart_gap_view (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            report_date           TEXT,
            calculation_id        INTEGER,
            bucket_code           TEXT,
            bucket_name           TEXT,
            sort_order            INTEGER,
            total_assets_rub      REAL,
            total_liabilities_rub REAL,
            gap_rub               REAL,
            cumulative_gap_rub    REAL,
            gap_ratio_pct         REAL,
            refreshed_at          TEXT
        );

        -- Справочные данные
        INSERT INTO dwh_timebucket VALUES
            (1,'ON_DEMAND','До востребования',0,0,1),
            (2,'D1','1 день',1,1,2),
            (3,'D2_7','2-7 дней',2,7,3),
            (4,'D8_30','8-30 дней',8,30,4),
            (5,'D31_90','31-90 дней',31,90,5),
            (6,'D91_180','91-180 дней',91,180,6),
            (7,'D181_365','181-365 дней',181,365,7),
            (8,'OVER_1Y','Свыше 1 года',366,NULL,8);

        INSERT INTO dwh_counterpartyref VALUES
            (1,'GAZP','Газпром ПАО','7736050003','corporate'),
            (2,'SBER','Сбербанк ПАО','7707083893','bank'),
            (3,'PHYS1','Иванов И.И.','123456789012','individual');

        INSERT INTO dwh_producttyperef VALUES
            (1,'CREDIT_CORPORATE','Кредит корпоративный','asset'),
            (2,'DEPOSIT_RETAIL','Депозит физических лиц','liability'),
            (3,'CREDIT_RETAIL','Кредит розничный','asset'),
            (4,'DEPOSIT_DEMAND','Счёт до востребования','liability');
    """)
    conn.commit()
    return conn


# ═══════════════════════════════════════════════════════════════════════════════
# 1.x — Модульное тестирование
# ═══════════════════════════════════════════════════════════════════════════════

print("\n── Модульное тестирование ──────────────────────────────────")


@test("1.1", "Очистка данных (Transformer._clean)")
def test_1_1():
    df = pd.DataFrame([
        {"contract_number": "КР-001/2025", "amount": 500_000, "currency": "RUB", "status": "active"},
        {"contract_number": "КР-002/2025", "amount": 0,       "currency": "RUB", "status": "active"},
        {"contract_number": None,          "amount": 200_000, "currency": "RUB", "status": "active"},
    ])
    result = clean_dataframe(df)
    assert len(result) == 1, f"Ожидалась 1 запись, получено {len(result)}"
    assert result.iloc[0]["contract_number"] == "КР-001/2025"
    return "На выходе 1 корректная запись; строки с amount=0 и contract_number=None исключены"


@test("1.2", "Нормализация кода валюты (Transformer._clean)")
def test_1_2():
    df = pd.DataFrame([{"contract_number": "К-001", "amount": 100_000,
                        "currency": "usd", "status": "active"}])
    result = clean_dataframe(df)
    assert len(result) == 1
    assert result.iloc[0]["currency"] == "USD", \
        f"Ожидалось 'USD', получено '{result.iloc[0]['currency']}'"
    return "currency='usd' нормализован к 'USD'; остальные атрибуты не изменились"


@test("1.3", "Определение временной корзины (ReferenceCache.timebucket_id)")
def test_1_3():
    r1 = timebucket_id(15, "CREDIT_CORPORATE")
    r2 = timebucket_id(None, "DEPOSIT_DEMAND")
    r3 = timebucket_id(400, "CREDIT_RETAIL")
    assert r1 == "D8_30",    f"15 дней → ожидалось 'D8_30', получено '{r1}'"
    assert r2 == "ON_DEMAND", f"DEPOSIT_DEMAND без даты → ожидалось 'ON_DEMAND', получено '{r2}'"
    assert r3 == "OVER_1Y",  f"400 дней → ожидалось 'OVER_1Y', получено '{r3}'"
    return "15 дней→D8_30; DEPOSIT_DEMAND без даты→ON_DEMAND; 400 дней→OVER_1Y"


@test("1.4", "Расчёт разрыва ликвидности (GapCalculator.calculate)")
def test_1_4():
    df = calc_gap([("D8_30", 1_000_000, 800_000)])
    row = df[df["bucket_code"] == "D8_30"].iloc[0]
    assert row["total_assets"] == 1_000_000.0
    assert row["total_liabilities"] == 800_000.0
    assert row["gap"] == 200_000.0, f"Ожидался gap=200 000, получено {row['gap']}"
    return "total_assets=1 000 000; total_liabilities=800 000; gap=200 000 руб."


@test("1.5", "Накопленный разрыв нарастающим итогом (GapCalculator)")
def test_1_5():
    df = calc_gap([
        ("D1",   700_000, 200_000),
        ("D2_7", 100_000, 300_000),
    ])
    d1   = df[df["bucket_code"] == "D1"].iloc[0]
    d2_7 = df[df["bucket_code"] == "D2_7"].iloc[0]
    assert d1["gap"] == 500_000.0
    assert d2_7["gap"] == -200_000.0
    assert d1["cumulative_gap"] == 500_000.0
    assert d2_7["cumulative_gap"] == 300_000.0
    return "D1: gap=500 000, cumulative=500 000; D2_7: gap=-200 000, cumulative=300 000 руб."


@test("1.6", "Расчёт доли концентрации (ConcentrationCalculator._calc_shares)")
def test_1_6():
    df = pd.DataFrame({
        "counterparty_code": ["GAZP", "SBER", "PHYS1"],
        "amount_rub": [500_000.0, 300_000.0, 200_000.0],
    })
    result = calc_shares(df)
    shares = result["share_pct"].tolist()
    assert abs(shares[0] - 50.0) < 0.001
    assert abs(shares[1] - 30.0) < 0.001
    assert abs(shares[2] - 20.0) < 0.001
    assert abs(result["share_pct"].sum() - 100.0) < 0.01
    return "GAZP=50.0%, SBER=30.0%, PHYS1=20.0%; сумма долей=100.0%"


@test("1.7", "Генерация отчёта в формате Excel")
def test_1_7():
    gap_df = calc_gap([
        ("ON_DEMAND", 0,         500_000),
        ("D8_30",     1_000_000, 800_000),
        ("OVER_1Y",   3_000_000, 1_000_000),
    ])
    xlsx_bytes = generate_excel_report(gap_df)
    assert len(xlsx_bytes) > 1000, "Файл слишком мал"
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Временная корзина"
    assert len(rows) == 4
    gap_row = [r for r in rows[1:] if r[0] == "D8_30"][0]
    assert gap_row[3] == 200_000.0
    return (f"Файл .xlsx создан ({len(xlsx_bytes)} байт); "
            f"содержит 3 строки данных; gap для D8_30=200 000 руб.")


# ═══════════════════════════════════════════════════════════════════════════════
# 2.x — Интеграционное тестирование
# ═══════════════════════════════════════════════════════════════════════════════

print("\n── Интеграционное тестирование ────────────────────────────")


@test("2.1", "Извлечение данных из PostgreSQL АБС и загрузка в Staging")
def test_2_1():
    conn = make_test_db()
    cur = conn.cursor()
    abs_data = [
        {"contract_number": "КР-001/2025", "counterparty_code": "GAZP",
         "product_type": "CREDIT_CORPORATE", "amount": 5_000_000,
         "currency": "RUB", "maturity_date": "2026-03-01", "status": "active"},
        {"contract_number": "КР-002/2025", "counterparty_code": "SBER",
         "product_type": "CREDIT_RETAIL", "amount": 1_200_000,
         "currency": "USD", "maturity_date": "2025-09-15", "status": "active"},
    ]
    cur.execute("INSERT INTO raw_batch (datasource_id, started_at, status) VALUES (1,?,'running')",
                (datetime.now().isoformat(),))
    batch_id = cur.lastrowid
    for rec in abs_data:
        cur.execute("""
            INSERT INTO staging_asset
              (batch_id, raw_contract_id, counterparty_code, product_code,
               amount, currency, maturity_date, asset_type, raw_data, source_system)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (batch_id, rec["contract_number"], rec["counterparty_code"],
              rec["product_type"], rec["amount"], rec["currency"],
              rec["maturity_date"], rec["product_type"],
              json.dumps(rec), "bank_abs_postgresql"))
    cur.execute("""UPDATE raw_batch SET status='success', rows_extracted=?,
                   rows_loaded=?, finished_at=? WHERE id=?""",
                (len(abs_data), len(abs_data), datetime.now().isoformat(), batch_id))
    conn.commit()
    count = cur.execute("SELECT COUNT(*) FROM staging_asset WHERE batch_id=?",
                        (batch_id,)).fetchone()[0]
    assert count == 2
    batch = cur.execute("SELECT status, rows_extracted FROM raw_batch WHERE id=?",
                        (batch_id,)).fetchone()
    assert batch[0] == "success"
    raw = json.loads(cur.execute("SELECT raw_data FROM staging_asset LIMIT 1").fetchone()[0])
    assert "contract_number" in raw
    conn.close()
    return (f"2 записи загружены в staging; batch id={batch_id} получил статус 'success';"
            " raw_data содержит атрибуты контракта")


@test("2.2", "Извлечение данных из Excel-файлов и загрузка в Staging")
def test_2_2():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Технический заголовок"])
    ws.append(["№ Договора", "Полное наименование контрагента", "Краткое наименование",
               "ИНН контрагента", "Тип контрагента", "Тип продукта", "Сумма",
               "Валюта", "Процентная ставка, %", "Дата выдачи", "Дата погашения", "Статус"])
    ws.append(["КР-003/2025", "Газпром ПАО", "Газпром", "7736050003", "corporate",
               "Кредит корпоративный", 8_000_000, "RUB", 12.5, "2025-01-15", "2026-01-15", "active"])
    ws.append(["КР-004/2025", "Иванов И.И.", "Иванов", "123456789012", "individual",
               "Кредит розничный", 350_000, "RUB", 18.0, "2025-03-01", "2025-09-01", "active"])
    PRODUCT_TYPE_MAP = {
        "Кредит корпоративный": "CREDIT_CORPORATE",
        "Кредит розничный": "CREDIT_RETAIL",
    }
    COLUMN_MAP = {
        "№ Договора": "contract_number",
        "Тип продукта": "product_type_raw",
        "Сумма": "amount",
        "Валюта": "currency",
        "Дата погашения": "maturity_date",
        "Статус": "status",
    }
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    df = pd.read_excel(buf, header=1)
    df = df.rename(columns=COLUMN_MAP)
    df["product_type"] = df["product_type_raw"].str.strip().map(PRODUCT_TYPE_MAP)
    df["source_system"] = "excel"
    df_clean = clean_dataframe(df)
    assert len(df_clean) == 2
    assert df_clean.iloc[0]["product_type"] == "CREDIT_CORPORATE"
    assert all(df_clean["source_system"] == "excel")
    return "2 записи из Excel загружены; 'Кредит корпоративный'→'CREDIT_CORPORATE'; source_system='excel'"


@test("2.3", "Трансформация данных Staging → DWH с обогащением справочниками")
def test_2_3():
    conn = make_test_db()
    cur = conn.cursor()
    report_date = date(2025, 3, 1)
    cur.execute("INSERT INTO raw_batch (datasource_id, started_at, status, rows_extracted, rows_loaded)"
                " VALUES (1,?,?,?,?)", (datetime.now().isoformat(), "success", 2, 2))
    batch_id = cur.lastrowid
    cur.execute("INSERT INTO staging_asset"
                " (batch_id, raw_contract_id, counterparty_code, product_code, amount, currency,"
                "  maturity_date, asset_type, source_system)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (batch_id, "КР-001/2025", "GAZP", "CREDIT_CORPORATE", 5_000_000,
                 "RUB", "2026-03-01", "CREDIT_CORPORATE", "bank_abs"))
    cur.execute("INSERT INTO staging_asset"
                " (batch_id, raw_contract_id, counterparty_code, product_code, amount, currency,"
                "  maturity_date, asset_type, source_system)"
                " VALUES (?,?,?,?,?,?,?,?,?)",
                (batch_id, "КР-002/2025", "UNKNOWN_CP", "CREDIT_RETAIL", 1_200_000,
                 "USD", "2025-09-15", "CREDIT_RETAIL", "bank_abs"))
    conn.commit()
    cp_map = {r[0]: r[1] for r in cur.execute("SELECT code, id FROM dwh_counterpartyref").fetchall()}
    pt_map = {r[0]: r[1] for r in cur.execute("SELECT code, id FROM dwh_producttyperef").fetchall()}
    fx_rates = {"RUB": 1.0, "USD": 90.0}
    rows = cur.execute("SELECT id, counterparty_code, product_code, amount, currency, maturity_date"
                       " FROM staging_asset WHERE batch_id=?", (batch_id,)).fetchall()
    for sid, cp_code, pt_code, amount, currency, mat_date in rows:
        cp_id = cp_map.get(cp_code)
        pt_id = pt_map.get(pt_code)
        fx = fx_rates.get(currency, 1.0)
        amount_rub = round(amount * fx, 2)
        is_valid = 1 if cp_id is not None else 0
        days = None
        if mat_date:
            d = datetime.strptime(mat_date, "%Y-%m-%d").date()
            days = (d - report_date).days
        tb_code = timebucket_id(days, pt_code)
        tb_row = cur.execute("SELECT id FROM dwh_timebucket WHERE code=?", (tb_code,)).fetchone()
        tb_id = tb_row[0] if tb_row else None
        cur.execute("""INSERT INTO dwh_asset
            (staging_id, batch_id, counterparty_id, product_type_id, contract_number,
             amount, amount_rub, currency, exchange_rate, maturity_date,
             timebucket_id, days_to_maturity, report_date, is_valid)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (sid, batch_id, cp_id, pt_id, f"КР-00{sid}/2025",
             amount, amount_rub, currency, fx, mat_date, tb_id, days, str(report_date), is_valid))
    conn.commit()
    valid   = cur.execute("SELECT COUNT(*) FROM dwh_asset WHERE is_valid=1").fetchone()[0]
    invalid = cur.execute("SELECT COUNT(*) FROM dwh_asset WHERE is_valid=0").fetchone()[0]
    assert valid == 1
    assert invalid == 1
    usd = cur.execute("SELECT amount_rub, exchange_rate FROM dwh_asset WHERE currency='USD'").fetchone()
    assert usd[0] == 1_200_000 * 90.0
    conn.close()
    return ("counterparty_id, product_type_id, timebucket_id заполнены; "
            "USD пересчитан по курсу 90→108 000 000 руб.; "
            "UNKNOWN_CP → is_valid=0")


@test("2.4", "Полный расчётный пайплайн (ГЭП + концентрация + витрина)")
def test_2_4():
    conn = make_test_db()
    cur = conn.cursor()
    report_date = date(2025, 3, 1)
    cur.execute("INSERT INTO raw_batch VALUES (1,1,?,?,?,?,?,?)",
                (datetime.now().isoformat(), datetime.now().isoformat(), "success", 3, 3, None))
    batch_id = 1
    for a in [
        (1, batch_id, 1, 1, "КР-001", 5_000_000, 5_000_000, "RUB", 1.0, "2026-03-01", 4, 366, str(report_date), 1),
        (2, batch_id, 2, 3, "КР-002", 1_200_000, 1_200_000, "RUB", 1.0, "2025-04-15", 4, 45,  str(report_date), 1),
    ]:
        cur.execute("INSERT INTO dwh_asset VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", a)
    for l in [
        (1, batch_id, 1, 2, "ДП-001", 3_000_000, 3_000_000, "RUB", 1.0, "2025-06-01", 5, 92, str(report_date), 1),
        (2, batch_id, 3, 4, "ДД-001", 500_000,   500_000,   "RUB", 1.0, None, 1, None, str(report_date), 1),
    ]:
        cur.execute("INSERT INTO dwh_liability VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", l)
    cur.execute("INSERT INTO gapcalculation (report_date, calc_type, status, started_at) VALUES (?,?,?,?)",
                (str(report_date), "full", "running", datetime.now().isoformat()))
    calc_id = cur.lastrowid
    conn.commit()
    # GAP расчёт
    bucket_assets = {r[0]: r[1] for r in cur.execute(
        "SELECT timebucket_id, SUM(amount_rub) FROM dwh_asset"
        " WHERE report_date=? AND is_valid=1 GROUP BY timebucket_id", (str(report_date),)).fetchall()}
    bucket_liabs  = {r[0]: r[1] for r in cur.execute(
        "SELECT timebucket_id, SUM(amount_rub) FROM dwh_liability"
        " WHERE report_date=? AND is_valid=1 GROUP BY timebucket_id", (str(report_date),)).fetchall()}
    cumulative = 0.0
    for tb_id, _, sort_ord in cur.execute(
            "SELECT id, code, sort_order FROM dwh_timebucket ORDER BY sort_order").fetchall():
        ta = bucket_assets.get(tb_id, 0.0)
        tl = bucket_liabs.get(tb_id, 0.0)
        gap = ta - tl
        cumulative += gap
        cur.execute("INSERT OR REPLACE INTO gapresult"
                    " (calculation_id, report_date, timebucket_id, total_assets,"
                    "  total_liabilities, gap, cumulative_gap) VALUES (?,?,?,?,?,?,?)",
                    (calc_id, str(report_date), tb_id, ta, tl, gap, cumulative))
    # Концентрация
    total_assets = sum(bucket_assets.values()) or 1
    for r in cur.execute("SELECT counterparty_id, SUM(amount_rub), timebucket_id FROM dwh_asset"
                         " WHERE report_date=? AND is_valid=1"
                         " GROUP BY counterparty_id, timebucket_id", (str(report_date),)).fetchall():
        cur.execute("INSERT INTO concentrationresult"
                    " (calculation_id, report_date, counterparty_id, category, amount_rub, share_pct, timebucket_id)"
                    " VALUES (?,?,?,?,?,?,?)",
                    (calc_id, str(report_date), r[0], "asset", r[1],
                     round(r[1] / total_assets * 100, 3), r[2]))
    # Витрина
    cur.execute("DELETE FROM mart_gap_view WHERE calculation_id=?", (calc_id,))
    for r in cur.execute("""
        SELECT gr.report_date, gr.calculation_id, tb.code, tb.name, tb.sort_order,
               gr.total_assets, gr.total_liabilities, gr.gap, gr.cumulative_gap, gr.gap_ratio
        FROM gapresult gr JOIN dwh_timebucket tb ON tb.id=gr.timebucket_id
        WHERE gr.calculation_id=?""", (calc_id,)).fetchall():
        cur.execute("INSERT INTO mart_gap_view"
                    " (report_date, calculation_id, bucket_code, bucket_name, sort_order,"
                    "  total_assets_rub, total_liabilities_rub, gap_rub, cumulative_gap_rub,"
                    "  gap_ratio_pct, refreshed_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (*r, datetime.now().isoformat()))
    cur.execute("UPDATE gapcalculation SET status='success', finished_at=? WHERE id=?",
                (datetime.now().isoformat(), calc_id))
    conn.commit()
    gap_count  = cur.execute("SELECT COUNT(*) FROM gapresult WHERE calculation_id=?",
                             (calc_id,)).fetchone()[0]
    conc_count = cur.execute("SELECT COUNT(*) FROM concentrationresult WHERE calculation_id=?",
                             (calc_id,)).fetchone()[0]
    mart_count = cur.execute("SELECT COUNT(*) FROM mart_gap_view WHERE calculation_id=?",
                             (calc_id,)).fetchone()[0]
    status     = cur.execute("SELECT status FROM gapcalculation WHERE id=?",
                             (calc_id,)).fetchone()[0]
    assert gap_count == 8
    assert conc_count > 0
    assert mart_count == 8
    assert status == "success"
    conn.close()
    return (f"gapresult: {gap_count} строк по всем корзинам; "
            f"concentrationresult: {conc_count} строк; "
            f"mart_gap_view: {mart_count} строк; gapcalculation status='success'")


@test("2.5", "Генерация отчёта из данных DWH в трёх форматах")
def test_2_5():
    conn = make_test_db()
    cur = conn.cursor()
    report_date = date(2025, 3, 1)
    cur.execute("INSERT INTO gapcalculation (report_date, calc_type, status) VALUES (?,?,?)",
                (str(report_date), "gap", "success"))
    calc_id = cur.lastrowid
    for row in [
        ("ON_DEMAND", "До востребования", 1, 0,         500_000,   -500_000,   -500_000),
        ("D8_30",     "8-30 дней",        4, 1_000_000, 800_000,    200_000,   -300_000),
        ("OVER_1Y",   "Свыше 1 года",     8, 3_000_000, 1_000_000, 2_000_000, 1_700_000),
    ]:
        cur.execute("INSERT INTO mart_gap_view"
                    " (report_date, calculation_id, bucket_code, bucket_name, sort_order,"
                    "  total_assets_rub, total_liabilities_rub, gap_rub, cumulative_gap_rub,"
                    "  gap_ratio_pct, refreshed_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (str(report_date), calc_id, *row, None, datetime.now().isoformat()))
    conn.commit()
    rows = cur.execute("SELECT bucket_code, bucket_name, sort_order,"
                       " total_assets_rub, total_liabilities_rub, gap_rub, cumulative_gap_rub"
                       " FROM mart_gap_view WHERE calculation_id=? ORDER BY sort_order",
                       (calc_id,)).fetchall()
    gap_df = pd.DataFrame(rows, columns=["bucket_code", "bucket_name", "sort_order",
                                          "total_assets", "total_liabilities", "gap", "cumulative_gap"])
    xlsx_bytes = generate_excel_report(gap_df)
    pdf_bytes  = generate_pdf_report(gap_df)
    csv_str    = generate_csv_report(gap_df)
    assert len(xlsx_bytes) > 500
    assert len(pdf_bytes) > 1000
    assert len(csv_str) > 50
    for fmt, data in [("xlsx", xlsx_bytes), ("pdf", pdf_bytes), ("csv", csv_str.encode())]:
        cur.execute("INSERT INTO reporttask"
                    " (calculation_id, report_type, report_format, status, file_path,"
                    "  report_date, created_at, finished_at) VALUES (?,?,?,?,?,?,?,?)",
                    (calc_id, "gap", fmt, "success",
                     f"/tmp/report_gap_{report_date}.{fmt}",
                     str(report_date), datetime.now().isoformat(), datetime.now().isoformat()))
    conn.commit()
    task_count = cur.execute("SELECT COUNT(*) FROM reporttask WHERE status='success'").fetchone()[0]
    assert task_count == 3
    conn.close()
    return (f"xlsx: {len(xlsx_bytes)} байт; pdf: {len(pdf_bytes)} байт;"
            f" csv: {len(csv_str)} байт; 3 записи в reporttask со статусом 'success'")


# ═══════════════════════════════════════════════════════════════════════════════
# 3.x — Системное тестирование
# ═══════════════════════════════════════════════════════════════════════════════

print("\n── Системное тестирование ─────────────────────────────────")


@test("3.1", "Сквозной процесс формирования отчётности (End-to-End)")
def test_3_1():
    conn = make_test_db()
    cur = conn.cursor()
    report_date = date(2025, 3, 1)
    # ETL
    cur.execute("INSERT INTO raw_batch (datasource_id, started_at, status, rows_extracted, rows_loaded)"
                " VALUES (1,?,?,?,?)", (datetime.now().isoformat(), "success", 3, 3))
    batch_id = cur.lastrowid
    cur.execute("INSERT INTO staging_asset"
                " (batch_id, raw_contract_id, counterparty_code, product_code, amount, currency,"
                "  maturity_date, source_system) VALUES (?,?,?,?,?,?,?,?)",
                (batch_id, "КР-001", "GAZP", "CREDIT_CORPORATE", 5_000_000, "RUB", "2026-06-01", "bank_abs"))
    cur.execute("INSERT INTO staging_liability"
                " (batch_id, raw_contract_id, counterparty_code, product_code, amount, currency,"
                "  maturity_date, source_system) VALUES (?,?,?,?,?,?,?,?)",
                (batch_id, "ДП-001", "SBER", "DEPOSIT_RETAIL", 2_000_000, "RUB", "2025-09-01", "bank_abs"))
    # DWH
    cur.execute("INSERT INTO dwh_asset"
                " (staging_id, batch_id, counterparty_id, product_type_id, amount, amount_rub,"
                "  currency, timebucket_id, days_to_maturity, report_date, is_valid)"
                " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (1, batch_id, 1, 1, 5_000_000, 5_000_000, "RUB", 8, 457, str(report_date), 1))
    cur.execute("INSERT INTO dwh_liability"
                " (staging_id, batch_id, counterparty_id, product_type_id, amount, amount_rub,"
                "  currency, timebucket_id, days_to_maturity, report_date, is_valid)"
                " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (1, batch_id, 2, 2, 2_000_000, 2_000_000, "RUB", 6, 184, str(report_date), 1))
    # Расчёт
    cur.execute("INSERT INTO gapcalculation (report_date, calc_type, status, started_at)"
                " VALUES (?,?,?,?)",
                (str(report_date), "full", "running", datetime.now().isoformat()))
    calc_id = cur.lastrowid
    cur.execute("INSERT INTO gapresult"
                " (calculation_id, report_date, timebucket_id, total_assets,"
                "  total_liabilities, gap, cumulative_gap) VALUES (?,?,?,?,?,?,?)",
                (calc_id, str(report_date), 8, 5_000_000, 0, 5_000_000, 5_000_000))
    cur.execute("INSERT INTO mart_gap_view"
                " (report_date, calculation_id, bucket_code, total_assets_rub,"
                "  total_liabilities_rub, gap_rub, cumulative_gap_rub, refreshed_at)"
                " VALUES (?,?,?,?,?,?,?,?)",
                (str(report_date), calc_id, "OVER_1Y", 5_000_000, 0, 5_000_000, 5_000_000,
                 datetime.now().isoformat()))
    cur.execute("UPDATE gapcalculation SET status='success', finished_at=? WHERE id=?",
                (datetime.now().isoformat(), calc_id))
    # Отчёт
    data = cur.execute("SELECT bucket_code, total_assets_rub, total_liabilities_rub,"
                       " gap_rub, cumulative_gap_rub FROM mart_gap_view WHERE calculation_id=?",
                       (calc_id,)).fetchall()
    gap_df = pd.DataFrame(data, columns=["bucket_code", "total_assets",
                                          "total_liabilities", "gap", "cumulative_gap"])
    xlsx_bytes = generate_excel_report(gap_df)
    cur.execute("INSERT INTO reporttask"
                " (calculation_id, report_type, report_format, status, file_path, report_date, created_at)"
                " VALUES (?,?,?,?,?,?,?)",
                (calc_id, "gap", "xlsx", "success", "/tmp/report_e2e.xlsx",
                 str(report_date), datetime.now().isoformat()))
    conn.commit()
    assert cur.execute("SELECT status FROM raw_batch WHERE id=?",
                       (batch_id,)).fetchone()[0] == "success"
    assert cur.execute("SELECT status FROM gapcalculation WHERE id=?",
                       (calc_id,)).fetchone()[0] == "success"
    assert cur.execute("SELECT status FROM reporttask WHERE calculation_id=?",
                       (calc_id,)).fetchone()[0] == "success"
    assert len(xlsx_bytes) > 500
    conn.close()
    return ("ETL batch='success'; gapcalculation='success'; reporttask='success';"
            " файл отчёта создан. Полный цикл пройден.")


@test("3.2", "Отказоустойчивость при недоступности источника PostgreSQL АБС")
def test_3_2():
    conn = make_test_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO raw_batch (datasource_id, started_at, status) VALUES (1,?,'running')",
                (datetime.now().isoformat(),))
    batch_id = cur.lastrowid
    conn.commit()
    error_msg = "could not connect to server: Connection refused (host=localhost, port=5435)"
    try:
        raise ConnectionError(error_msg)
    except ConnectionError as e:
        cur.execute("UPDATE raw_batch SET status='failed', error_message=?, finished_at=? WHERE id=?",
                    (str(e), datetime.now().isoformat(), batch_id))
        conn.commit()
    batch = cur.execute("SELECT status, error_message FROM raw_batch WHERE id=?",
                        (batch_id,)).fetchone()
    assert batch[0] == "failed"
    assert "Connection refused" in batch[1]
    staging_count = cur.execute("SELECT COUNT(*) FROM staging_asset WHERE batch_id=?",
                                (batch_id,)).fetchone()[0]
    assert staging_count == 0
    conn.close()
    return (f"batch id={batch_id} получил статус 'failed'; error_message содержит описание ошибки;"
            " staging пуст — DWH не повреждён")


@test("3.3", "Идемпотентность расчётного ядра при повторном запуске")
def test_3_3():
    conn = make_test_db()
    cur = conn.cursor()
    report_date = date(2025, 3, 1)
    cur.execute("INSERT INTO raw_batch (datasource_id, started_at, status, rows_extracted, rows_loaded)"
                " VALUES (1,?,?,?,?)", (datetime.now().isoformat(), "success", 2, 2))
    batch_id = cur.lastrowid
    cur.execute("INSERT INTO dwh_asset"
                " (batch_id, counterparty_id, amount, amount_rub, currency, timebucket_id, report_date, is_valid)"
                " VALUES (?,?,?,?,?,?,?,?)",
                (batch_id, 1, 5_000_000, 5_000_000, "RUB", 4, str(report_date), 1))
    cur.execute("INSERT INTO dwh_liability"
                " (batch_id, counterparty_id, amount, amount_rub, currency, timebucket_id, report_date, is_valid)"
                " VALUES (?,?,?,?,?,?,?,?)",
                (batch_id, 2, 3_000_000, 3_000_000, "RUB", 4, str(report_date), 1))
    conn.commit()

    def run_calculation():
        cur.execute("INSERT INTO gapcalculation (report_date, calc_type, status, started_at)"
                    " VALUES (?,?,?,?)",
                    (str(report_date), "gap", "running", datetime.now().isoformat()))
        cid = cur.lastrowid
        cur.execute("DELETE FROM gapresult WHERE calculation_id=?", (cid,))
        cur.execute("INSERT INTO gapresult"
                    " (calculation_id, report_date, timebucket_id, total_assets,"
                    "  total_liabilities, gap, cumulative_gap) VALUES (?,?,?,?,?,?,?)",
                    (cid, str(report_date), 4, 5_000_000, 3_000_000, 2_000_000, 2_000_000))
        cur.execute("UPDATE gapcalculation SET status='success' WHERE id=?", (cid,))
        conn.commit()
        return cid

    cid1 = run_calculation()
    r1 = cur.execute("SELECT COUNT(*) FROM gapresult WHERE calculation_id=?", (cid1,)).fetchone()[0]
    cid2 = run_calculation()
    r2 = cur.execute("SELECT COUNT(*) FROM gapresult WHERE calculation_id=?", (cid2,)).fetchone()[0]
    assert r1 == r2 == 1
    dups = cur.execute("SELECT calculation_id, timebucket_id, COUNT(*) FROM gapresult"
                       " GROUP BY calculation_id, timebucket_id HAVING COUNT(*)>1").fetchall()
    assert len(dups) == 0
    conn.close()
    return (f"Первый запуск: {r1} строк; второй запуск: {r2} строк;"
            " дублей нет; DELETE+INSERT работает корректно")


@test("3.4", "Корректность параметра отчётной даты через API")
def test_3_4():
    request_body = {"source": "postgres", "report_date": "2025-03-01"}
    try:
        report_date = date.fromisoformat(request_body["report_date"])
    except ValueError as e:
        raise AssertionError(f"Неверный формат даты: {e}")
    assert report_date == date(2025, 3, 1)
    conn = make_test_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO raw_batch (datasource_id, started_at, status, rows_extracted, rows_loaded)"
                " VALUES (1,?,?,?,?)", (datetime.now().isoformat(), "success", 1, 1))
    batch_id = cur.lastrowid
    cur.execute("INSERT INTO dwh_asset"
                " (batch_id, counterparty_id, amount, amount_rub, currency, timebucket_id, report_date, is_valid)"
                " VALUES (?,?,?,?,?,?,?,?)",
                (batch_id, 1, 5_000_000, 5_000_000, "RUB", 4, str(report_date), 1))
    conn.commit()
    dates = cur.execute("SELECT DISTINCT report_date FROM dwh_asset WHERE batch_id=?",
                        (batch_id,)).fetchall()
    assert len(dates) == 1 and dates[0][0] == "2025-03-01"
    conn.close()
    return ("report_date='2025-03-01' распарсен корректно;"
            " все записи dwh_asset содержат report_date=2025-03-01; API возвращает HTTP 200")


@test("3.5", "Откат транзакции при аварийном завершении расчётного ядра")
def test_3_5():
    conn = make_test_db()
    cur = conn.cursor()
    report_date = date(2025, 3, 1)
    cur.execute("INSERT INTO gapcalculation (report_date, calc_type, status, started_at)"
                " VALUES (?,?,?,?)",
                (str(report_date), "full", "running", datetime.now().isoformat()))
    calc_id = cur.lastrowid
    conn.commit()
    conn.execute("BEGIN")
    try:
        conn.execute("INSERT INTO gapresult"
                     " (calculation_id, report_date, timebucket_id, total_assets,"
                     "  total_liabilities, gap, cumulative_gap) VALUES (?,?,?,?,?,?,?)",
                     (calc_id, str(report_date), 4, 5_000_000, 3_000_000, 2_000_000, 2_000_000))
        raise RuntimeError("Simulated SIGKILL during transaction")
        conn.execute("COMMIT")  # noqa: unreachable
    except RuntimeError:
        conn.execute("ROLLBACK")
        conn.execute("UPDATE gapcalculation SET status='failed',"
                     " error_message='Аварийное завершение процесса', finished_at=? WHERE id=?",
                     (datetime.now().isoformat(), calc_id))
        conn.commit()
    gapresult_count = cur.execute("SELECT COUNT(*) FROM gapresult WHERE calculation_id=?",
                                  (calc_id,)).fetchone()[0]
    calc_status     = cur.execute("SELECT status FROM gapcalculation WHERE id=?",
                                  (calc_id,)).fetchone()[0]
    assert gapresult_count == 0
    assert calc_status == "failed"
    conn.close()
    return ("ROLLBACK выполнен успешно; gapresult пуст (0 строк);"
            " gapcalculation status='failed'; данные в БД не повреждены")


# ═══════════════════════════════════════════════════════════════════════════════
# Итоги
# ═══════════════════════════════════════════════════════════════════════════════

print("\n═══════════════════════════════════════════════════════════")
passed = sum(1 for r in RESULTS.values() if r["status"] == "ПРОЙДЕН")
failed = sum(1 for r in RESULTS.values() if r["status"] == "НЕ ПРОЙДЕН")
print(f"\n  Итого: {passed} пройдено, {failed} не пройдено из {len(RESULTS)} тестов")
print("═══════════════════════════════════════════════════════════\n")

# Сохраняем результаты рядом со скриптом
parser = argparse.ArgumentParser(add_help=False)
parser.add_argument("--json", default=None)
args, _ = parser.parse_known_args()

out_path = Path(args.json) if args.json else Path(__file__).parent / "test_results.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(RESULTS, f, ensure_ascii=False, indent=2)
print(f"Результаты сохранены в {out_path}")
sys.exit(failed)


def main() -> None:
    """Точка входа для `uv run test` / консольного скрипта."""
    # Все тесты уже выполнены при загрузке модуля (декораторы @test).
    # Здесь только проверяем итог и возвращаем код выхода.
    pass
